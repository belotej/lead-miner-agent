import pandas as pd
import os
from openpyxl import Workbook, load_workbook

class ExcelWriter:
    def __init__(self, filepath):
        self.filepath = filepath
        self.sheets = ["Raw Discoveries", "Enriched", "Scored", "Ready for Outreach", "Historical"]
        self._initialize_workbook()

    def _initialize_workbook(self):
        if not os.path.exists(self.filepath):
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            for sheet_name in self.sheets:
                wb.create_sheet(sheet_name)
                
            # Add headers to Raw Discoveries
            raw_sheet = wb["Raw Discoveries"]
            headers = [
                "Discovery Date", "Company Name", "Domain", "Discovery Source", 
                "Signal Type", "Signal Strength", "Signal Date", "Details", 
                "Location", "Timeline", "Source URL", "County", "All Signals", "Notes"
            ]
            raw_sheet.append(headers)
            
            wb.save(self.filepath)
            print(f"Created new leads repository at {self.filepath}")

    def save_leads(self, leads, sheet_name="Raw Discoveries"):
        # leads is a list of dictionaries
        if not leads:
            return
            
        df_new = pd.DataFrame(leads)
        
        try:
            # Load existing data
            with pd.ExcelWriter(self.filepath, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                # We need to find the last row to append
                # For now, let's just read existing and append (inefficient for huge files but safe)
                # Better approach for openpyxl:
                pass
                
            # Using openpyxl directly for appending
            wb = load_workbook(self.filepath)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            
            for lead in leads:
                # Map dictionary to row based on headers
                # This assumes headers are in the first row
                # For simplicity in this step, we will just append values in a specific order matching headers
                # A more robust way is to map keys to column indices
                
                row = []
                # Based on the headers we defined
                headers = [cell.value for cell in ws[1]]
                for header in headers:
                    # Convert header to key (e.g. "Discovery Date" -> "discovery_date")
                    # But the incoming data might use different keys. 
                    # Let's assume the scraper output matches the headers or we map them.
                    # The plan has specific output schema.
                    # For now, let's just assume the dict keys match the header names for simplicity
                    # or we can do a simple mapping here.
                    
                    key = header.lower().replace(" ", "_")
                    row.append(lead.get(key, ""))
                
                ws.append(row)
                
            wb.save(self.filepath)
            print(f"Saved {len(leads)} leads to {sheet_name}")
            
        except Exception as e:
            print(f"Error saving leads: {e}")
